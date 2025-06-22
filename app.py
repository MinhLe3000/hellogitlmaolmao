from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
import os
from openpyxl import Workbook, load_workbook

app = FastAPI()

# Thêm middleware CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Cho phép tất cả nguồn truy cập
    allow_credentials=True,
    allow_methods=["*"],  # Cho phép tất cả các phương thức
    allow_headers=["*"],  # Cho phép tất cả các header
)

# Đường dẫn đến file lưu cơ sở tri thức dưới dạng Excel
file_path = 'knowledgebase.xlsx'

# Danh sách phòng ban cố định
FIXED_DEPARTMENTS = {
    1: "Trung tâm Ứng dụng CNTT",
    2: "Trung tâm Dịch vụ Hỗ trợ Đào tạo",
    3: "Trung tâm Thông tin - Thư viện",
    4: "Phòng Truyền thông",
    5: "Phòng Đào tạo",
    6: "Phòng Công tác CTCT & QLSV"
}
OTHER_DEPARTMENT = "Other Departments"

# Tải cơ sở tri thức
def load_knowledge_base():
    knowledge_base = {}
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active

        departments = [cell.value for cell in sheet[1] if cell.value]
        for col, department in enumerate(departments, start=1):
            for row in sheet.iter_rows(min_row=2, min_col=col, max_col=col, values_only=True):
                keyword = row[0]
                if keyword:
                    knowledge_base[keyword.strip().lower()] = department.strip()
    return knowledge_base

# Lưu cơ sở tri thức
def save_knowledge_base(knowledge_base):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()

    sheet = workbook.active
    departments = list(FIXED_DEPARTMENTS.values()) + [OTHER_DEPARTMENT]
    for idx, department in enumerate(departments, start=1):
        sheet.cell(row=1, column=idx, value=department)

    department_columns = {department: idx + 1 for idx, department in enumerate(departments)}
    for col in range(1, len(departments) + 1):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=col, value=None)

    department_data = {dept: [] for dept in departments}
    for keyword, department in knowledge_base.items():
        department_data[department].append(keyword)

    for department, keywords in department_data.items():
        col = department_columns[department]
        for row, keyword in enumerate(keywords, start=2):
            sheet.cell(row=row, column=col, value=keyword)

    workbook.save(file_path)

# Cơ sở tri thức toàn cục
knowledge_base = load_knowledge_base()

# Mô hình dữ liệu cho API
class KeywordRequest(BaseModel):
    keyword: str
    department_id: int

# API để tìm kiếm phòng ban
@app.get("/find-department/")
def find_department(keyword: str):
    keyword = keyword.lower()
    department = knowledge_base.get(keyword)
    if department:
        return {"keyword": keyword, "department": department}
    else:
        raise HTTPException(status_code=404, detail="Keyword not found")

# API để thêm từ khóa
@app.post("/add-keyword/")
def add_to_knowledge_base(data: KeywordRequest):
    keyword = data.keyword.lower()
    if data.department_id in FIXED_DEPARTMENTS:
        department = FIXED_DEPARTMENTS[data.department_id]
    else:
        department = OTHER_DEPARTMENT

    if keyword not in knowledge_base:
        knowledge_base[keyword] = department
        save_knowledge_base(knowledge_base)
        return {"message": f"Keyword '{keyword}' added to department '{department}'."}
    else:
        return {"message": f"Keyword '{keyword}' already exists in department '{knowledge_base[keyword]}'."}

# API kiểm tra tất cả từ khóa
@app.get("/list-keywords/")
def list_keywords():
    return {"knowledge_base": knowledge_base}

# API để lấy danh sách phòng ban
@app.get("/list-departments/")
def list_departments():
    return {"departments": FIXED_DEPARTMENTS, "other_department": OTHER_DEPARTMENT}